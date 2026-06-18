from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "powerbi_dashboard" / "data"
OUT = ROOT / "powerbi_dashboard" / "DataVis_PowerBI_Input.xlsx"


SHEETS = [
    ("README", None),
    ("KPISummary", DATA / "kpi_summary.csv"),
    ("OverallGroupSummary", DATA / "overall_group_summary_long.csv"),
    ("SignupTrend", DATA / "overall_signup_month_trend.csv"),
    ("RiskLevelSummary", DATA / "risk_level_summary.csv"),
    ("RiskByFeature", DATA / "risk_summary_by_feature.csv"),
    ("RiskTop5000", DATA / "risk_top_5000_customers.csv"),
    ("ModelMetrics", DATA / "model_metrics_flat.csv"),
    ("ConfusionMatrix", DATA / "model_confusion_matrix.csv"),
    ("SplitSummary", DATA / "model_split_summary.csv"),
    ("FeatureImportance", DATA / "model_feature_importance.csv"),
    ("ChurnOnlyProfile", DATA / "churn_only_profile_summary_long.csv"),
]


def coerce_value(value: str):
    if value == "":
        return None
    text = value.strip()
    if text.lower() in {"true", "false"}:
        return text.lower() == "true"
    try:
        if any(ch in text for ch in [".", "e", "E"]):
            return float(text)
        return int(text)
    except ValueError:
        return value


def autosize(ws, max_width: int = 42) -> None:
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            widths[cell.column] = min(max(widths.get(cell.column, 0), length + 2), max_width)
    for column_index, width in widths.items():
        ws.column_dimensions[get_column_letter(column_index)].width = width


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_csv_sheet(wb: Workbook, sheet_name: str, csv_path: Path) -> None:
    ws = wb.create_sheet(sheet_name)
    with csv_path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append([coerce_value(value) for value in row])
    style_header(ws)
    autosize(ws)


def write_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("README")
    rows = [
        ["DataVis Power BI Input Workbook"],
        [""],
        ["Use this file in Power BI Desktop: Get Data > Excel workbook."],
        ["Main dashboard version: overall."],
        ["Support page: churn_only, used only for profiling already churned customers."],
        [""],
        ["Recommended pages"],
        ["1. Overall Churn Overview"],
        ["2. Churn Drivers"],
        ["3. Prediction and Retention Priority"],
        ["4. Churn-Only Support"],
        ["5. Model Evaluation"],
        [""],
        ["Important wording"],
        ["For ChurnOnlyProfile, use 'Percentage of Churned Customers', not 'Churn Rate'."],
    ]
    for row in rows:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=16)
    autosize(ws, max_width=90)


def main() -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for sheet_name, csv_path in SHEETS:
        if sheet_name == "README":
            write_readme(wb)
        else:
            write_csv_sheet(wb, sheet_name, csv_path)

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
